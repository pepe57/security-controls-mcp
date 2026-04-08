#!/usr/bin/env python3
"""
Extract all frameworks from SCF spreadsheet (dynamic, version-agnostic).

This script parses the official SCF Excel file and generates:
1. scf-controls.json - All controls with complete framework mappings
2. framework-to-scf.json - Reverse index from framework controls to SCF IDs
3. framework-metadata.json - Framework lookup table with categories and stats

The script auto-detects framework columns from the spreadsheet headers instead
of hardcoding column indices, so it works across SCF version upgrades without
code changes.

Usage:
    poetry run python scripts/extract_scf_frameworks.py
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
SCF_XLSX = SCRIPT_DIR / "data" / "scf-2026.1.xlsx"
SCF_SHEET = "SCF 2026.1"
OUTPUT_DIR = PROJECT_ROOT / "src" / "security_controls_mcp" / "data"

# --- Metadata column indices (0-based) ---
# These are stable across SCF versions; verified against 2025.4 and 2026.1.
COL_DOMAIN = 0          # SCF Domain
COL_CONTROL_NAME = 1    # SCF Control
COL_SCF_ID = 2          # SCF #
COL_DESCRIPTION = 3     # Control Description
COL_CADENCE = 4         # Conformity Validation Cadence
COL_WEIGHT = 12         # Relative Control Weighting
COL_PPTDF = 13          # PPTDF Applicability

# --- Header patterns that mark non-framework columns ---
# Columns matching these prefixes are metadata, risk/threat mappings, or
# internal SCF fields -- not external frameworks. Checked against the raw
# header text (which may contain newlines).
NON_FRAMEWORK_PREFIXES = (
    "SCF",               # SCF Domain, SCF Control, SCF #, SCF CORE, SCF Community, etc.
    "Secure Controls",   # Full control description header
    "Conformity",        # Validation cadence
    "Evidence Request",  # ERL #
    "Possible Solutions",# Business-size recommendations
    "Relative Control",  # Weight
    "PPTDF",             # Applicability
    "NIST CSF\n",        # Function Grouping (metadata, not the NIST CSF framework mapping)
    "SCRM Focus",        # SCRM tier columns
    "SCR-CMM",           # Maturity model levels
    "Minimum Security",  # MCR + DSR summary
    "Identify\n",        # MCR / DSR identify columns
    "Risk",              # Risk factor columns
    "Threat",            # Threat catalog columns
    "Control Threat",    # Control-level threat summary
    "Errata",            # Errata column
)


def normalize_framework_id(header: str) -> str:
    """Convert a spreadsheet header into a stable framework_id.

    Rules:
    - Collapse newlines and multiple spaces into single space
    - Lowercase
    - Replace spaces, dashes, slashes, colons with underscores
    - Strip leading/trailing underscores
    - Collapse multiple underscores
    - Remove parentheses content where it's just noise, but keep
      identifiers like (SOC 2), (GDPR), etc. intact
    """
    # Normalize whitespace
    s = re.sub(r"[\n\r]+", " ", header)
    s = re.sub(r"\s+", " ", s).strip()

    # Lowercase
    s = s.lower()

    # Replace separators with underscores
    s = re.sub(r"[\s\-/:\\]+", "_", s)

    # Remove stray punctuation but keep dots (version numbers) and parens
    s = re.sub(r"[,;'\"&]+", "", s)

    # Collapse multiple underscores
    s = re.sub(r"_+", "_", s)

    # Strip leading/trailing underscores
    s = s.strip("_")

    return s


def build_display_name(header: str) -> str:
    """Clean up a spreadsheet header into a human-readable display name.

    Replaces newlines with spaces, collapses whitespace, strips.
    """
    s = re.sub(r"[\n\r]+", " ", header)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_framework_column(header: str) -> bool:
    """Return True if the column header represents an external framework mapping.

    Filters out metadata columns, SCF internal columns, and risk/threat columns.
    """
    if not header or not header.strip():
        return False

    raw = header.strip()

    for prefix in NON_FRAMEWORK_PREFIXES:
        if raw.startswith(prefix):
            return False

    return True


def detect_framework_columns(ws) -> list[tuple[int, str, str]]:
    """Auto-detect framework columns from the header row.

    Returns list of (column_index, framework_id, display_name).
    """
    frameworks = []
    seen_ids = {}

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for col_idx, cell_value in enumerate(row):
            header = str(cell_value).strip() if cell_value else ""
            if not is_framework_column(header):
                continue

            fw_id = normalize_framework_id(header)
            display = build_display_name(header)

            # Deduplicate: if two columns produce the same fw_id, append column index
            if fw_id in seen_ids:
                fw_id = f"{fw_id}_{col_idx}"
            seen_ids[fw_id] = col_idx

            frameworks.append((col_idx, fw_id, display))

    return frameworks


# --- Framework categories ---
# Assigns each framework_id to one or more browsing categories.
# Categories use keyword matching on the framework_id to auto-assign,
# with an explicit override map for edge cases. This keeps the mapping
# maintainable across SCF version changes: new frameworks with recognizable
# prefixes get categorized automatically.

CATEGORY_RULES: dict[str, list[str]] = {
    # Pattern-based: if framework_id contains any of these substrings, assign to category
    "ai_governance": ["42001", "ai_rmf", "ai_600", "ai_act", "ai_enabled", "ai_model"],
    "iso_standards": ["iso_"],
    "nist_frameworks": ["nist_"],
    "cis_controls": ["cis_csc", "cis_"],
    "cloud_security": ["27017", "27018", "csa_ccm", "csa_iot", "c5_2020", "c5_"],
    "governance": ["cobit", "coso", "cr_cmm"],
    "privacy": [
        "gdpr", "27701", "29100", "privacy", "apec_", "gapp", "oecd_",
        "dpf", "ccpa", "cpa", "cdpa", "coppa", "ferpa", "fipps",
        "bipa", "ipa", "pipa", "pdpl", "pdppl", "dpa_", "dpr_",
        "lgpd", "dpdpa", "pipl", "appi", "pipeda", "popia",
    ],
    "financial": [
        "soc_2", "tsc", "pci_dss", "sox", "glba", "ffiec", "finra",
        "sec_cybersecurity", "facta", "ftc_act", "naic", "fca_crm",
        "dora", "psd2", "eba_", "swift", "shared_assessments",
        "mas_trm", "sama_csf", "osfi", "cps_230", "cps_234",
        "sebi_cscrf", "bait",
    ],
    "healthcare": [
        "hipaa", "cms_mars", "hhs_", "fda_", "60601",
        "800_66",
    ],
    "industrial_ot": [
        "62443", "nerc_cip", "800_82", "maritime", "tsa_dhs",
    ],
    "automotive": [
        "21434", "tisax", "un_r155", "un_ece", "ul_2900",
    ],
    "supply_chain": [
        "800_161", "800_218", "dfars",
    ],
    "threat_intel_appsec": ["mitre", "owasp", "sparta"],
    "zero_trust": ["zero_trust", "zt_", "ztra", "ztcf", "800_207"],
    "fedramp": ["fedramp"],
    "govramp": ["govramp"],
    "cmmc": ["cmmc"],
    "us_federal": [
        "cjis", "irs_", "nispom", "nnpi", "eo_14028",
        "dhs_cisa", "c2m2", "cert_rmm", "cisa_cpg",
        "far_52", "itar", "nstc_", "ssa_eiesr",
    ],
    "us_state_laws": [
        "us_ak", "us_ca", "us_co", "us_il", "us_ma",
        "us_nv", "us_ny", "us_or", "us_tn", "us_tx",
        "us_va", "us_vt", "nydfs", "tx_ramp",
    ],
    "eu_regulations": [
        "eu_ai", "eu_cyber", "eu_eba", "dora", "gdpr",
        "nis2", "psd2", "enisa",
    ],
    "europe_national": [
        "emea_austria", "emea_belgium", "emea_germany", "germany",
        "emea_greece", "emea_hungary", "emea_ireland", "emea_italy",
        "emea_netherlands", "netherlands", "emea_norway", "norway",
        "emea_poland", "poland", "emea_sweden", "sweden",
        "emea_spain", "spain", "emea_switzerland", "switzerland",
        "emea_turkey", "turkey", "emea_russia", "russia",
        "emea_serbia", "serbia", "emea_uk", "uk_caf", "uk_cap",
        "uk_cyber", "uk_defstan", "uk_dpa", "fca_crm", "bsi_",
    ],
    "middle_east_africa": [
        "israel", "saudi", "uae_", "qatar", "south_africa",
        "kenya", "nigeria", "emea_south_africa",
    ],
    "asia_pacific": [
        "apac_", "australia", "singapore", "japan", "china",
        "hong_kong", "india", "south_korea", "taiwan",
        "malaysia", "philippines", "new_zealand", "nz_",
    ],
    "americas": [
        "americas_", "argentina", "bahamas", "bermuda",
        "brazil", "canada", "chile", "colombia", "mexico",
        "peru", "costa_rica", "uruguay",
    ],
    "media_entertainment": ["mpa_"],
}

# Explicit overrides: framework_id -> list of categories to ADD
CATEGORY_OVERRIDES: dict[str, list[str]] = {
    # Frameworks that need manual assignment because their ID doesn't match patterns
}


def assign_categories(framework_ids: list[str]) -> dict[str, list[str]]:
    """Build category -> [framework_id] mapping using pattern rules + overrides."""
    categories: dict[str, list[str]] = {cat: [] for cat in CATEGORY_RULES}

    for fw_id in framework_ids:
        assigned = set()
        for cat, patterns in CATEGORY_RULES.items():
            for pattern in patterns:
                if pattern in fw_id:
                    assigned.add(cat)
                    break

        # Apply explicit overrides
        if fw_id in CATEGORY_OVERRIDES:
            for cat in CATEGORY_OVERRIDES[fw_id]:
                assigned.add(cat)

        for cat in assigned:
            if cat in categories:
                categories[cat].append(fw_id)

    # Remove empty categories
    return {cat: sorted(fws) for cat, fws in categories.items() if fws}


def parse_cell_value(value) -> list[str]:
    """Parse a cell value into a list of control IDs."""
    if not value:
        return []

    # Convert to string and clean
    text = str(value).strip()
    if not text or text.lower() == "none":
        return []

    # Split by newlines and clean each line
    controls = []
    for line in text.split("\n"):
        line = line.strip()
        if line and line.lower() != "none":
            # Handle multiple IDs on same line (comma or semicolon separated)
            for part in re.split(r"[,;]", line):
                part = part.strip()
                if part:
                    controls.append(part)

    return controls


def extract_controls(ws, framework_columns: list[tuple[int, str, str]]) -> list[dict]:
    """Extract all controls from the worksheet."""
    controls = []
    row_count = 0

    # Build column index lookup
    fw_col_map = {col_idx: (fw_id, display) for col_idx, fw_id, display in framework_columns}

    print("Extracting controls...")

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[COL_SCF_ID]:
            continue

        scf_id = str(row[COL_SCF_ID]).strip()
        if not scf_id or scf_id == "SCF #":
            continue

        # Extract core fields
        weight_raw = row[COL_WEIGHT] if COL_WEIGHT < len(row) else None
        weight = 5  # default
        if weight_raw is not None:
            try:
                weight = int(weight_raw)
            except (ValueError, TypeError):
                weight = 5

        control = {
            "id": scf_id,
            "domain": str(row[COL_DOMAIN]).strip() if row[COL_DOMAIN] else "",
            "name": str(row[COL_DESCRIPTION]).strip() if row[COL_DESCRIPTION] else "",
            "description": str(row[COL_DESCRIPTION]).strip() if row[COL_DESCRIPTION] else "",
            "weight": weight,
            "pptdf": str(row[COL_PPTDF]).strip() if COL_PPTDF < len(row) and row[COL_PPTDF] else "",
            "validation_cadence": str(row[COL_CADENCE]).strip() if row[COL_CADENCE] else "Annual",
            "framework_mappings": {}
        }

        # Extract all framework mappings
        for col_idx, (fw_id, _display) in fw_col_map.items():
            try:
                cell_value = row[col_idx] if col_idx < len(row) else None
                mappings = parse_cell_value(cell_value)
                control["framework_mappings"][fw_id] = mappings if mappings else None
            except IndexError:
                control["framework_mappings"][fw_id] = None

        controls.append(control)
        row_count += 1

        if row_count % 100 == 0:
            print(f"  Processed {row_count} controls...")

    print(f"Extracted {len(controls)} controls")
    return controls


def build_reverse_index(
    controls: list[dict],
    framework_ids: list[str],
) -> dict[str, dict[str, list[str]]]:
    """Build reverse index from framework controls to SCF IDs."""
    reverse_index: dict[str, dict[str, list[str]]] = {fw_id: {} for fw_id in framework_ids}

    for control in controls:
        scf_id = control["id"]
        for fw_key, mappings in control["framework_mappings"].items():
            if mappings:
                for mapping in mappings:
                    if mapping not in reverse_index.get(fw_key, {}):
                        reverse_index[fw_key][mapping] = []
                    reverse_index[fw_key][mapping].append(scf_id)

    return reverse_index


def get_framework_stats(
    controls: list[dict],
    framework_ids: list[str],
) -> dict[str, int]:
    """Calculate control counts per framework."""
    stats = {}
    for fw_id in framework_ids:
        count = sum(1 for c in controls if c["framework_mappings"].get(fw_id))
        stats[fw_id] = count
    return stats


def main():
    print(f"Loading SCF spreadsheet: {SCF_XLSX}")
    print(f"Sheet: {SCF_SHEET}")
    wb = load_workbook(SCF_XLSX, data_only=True)
    ws = wb[SCF_SHEET]

    # Auto-detect framework columns
    print("Detecting framework columns...")
    framework_columns = detect_framework_columns(ws)
    framework_ids = [fw_id for _, fw_id, _ in framework_columns]
    print(f"  Found {len(framework_columns)} framework columns "
          f"(cols {framework_columns[0][0]}-{framework_columns[-1][0]})")

    # Extract controls
    controls = extract_controls(ws, framework_columns)
    wb.close()

    # Build reverse index
    print("Building reverse index...")
    reverse_index = build_reverse_index(controls, framework_ids)

    # Calculate stats
    stats = get_framework_stats(controls, framework_ids)

    # Build categories
    print("Assigning framework categories...")
    categories = assign_categories(framework_ids)
    categorized = set()
    for fws in categories.values():
        categorized.update(fws)
    uncategorized = [fw for fw in framework_ids if fw not in categorized]
    if uncategorized:
        print(f"  WARNING: {len(uncategorized)} uncategorized frameworks:")
        for fw in uncategorized:
            display = next(d for _, fid, d in framework_columns if fid == fw)
            print(f"    - {fw} ({display})")
    print(f"  {len(categories)} categories, {len(categorized)} frameworks categorized")

    # Output statistics
    print(f"\n=== FRAMEWORK STATISTICS ===")
    print(f"Total frameworks: {len(framework_columns)}")
    print(f"Total controls: {len(controls)}")

    # Sort by control count
    sorted_stats = sorted(stats.items(), key=lambda x: x[1], reverse=True)
    print("\nTop 20 frameworks by control count:")
    for fw_id, count in sorted_stats[:20]:
        display = next(d for _, fid, d in framework_columns if fid == fw_id)
        print(f"  {display}: {count} controls")

    # Tier 0 stats (AI governance)
    print("\n=== TIER 0 (AI GOVERNANCE) ===")
    tier0_patterns = ["42001", "ai_rmf", "ai_600", "ai_act"]
    for fw_id in framework_ids:
        if any(p in fw_id for p in tier0_patterns):
            count = stats.get(fw_id, 0)
            display = next(d for _, fid, d in framework_columns if fid == fw_id)
            print(f"  {display}: {count} controls")

    # Save outputs
    print(f"\nSaving to {OUTPUT_DIR}...")

    # Save controls
    controls_file = OUTPUT_DIR / "scf-controls.json"
    with open(controls_file, "w", encoding="utf-8") as f:
        json.dump({"controls": controls}, f, indent=2, ensure_ascii=False)
    print(f"  Saved {controls_file}")

    # Save reverse index
    reverse_file = OUTPUT_DIR / "framework-to-scf.json"
    with open(reverse_file, "w", encoding="utf-8") as f:
        json.dump(reverse_index, f, indent=2, ensure_ascii=False)
    print(f"  Saved {reverse_file}")

    # Save framework metadata
    metadata_file = SCRIPT_DIR / "data" / "framework-metadata.json"
    fw_lookup = {fw_id: {"id": fw_id, "name": display} for _, fw_id, display in framework_columns}
    metadata = {
        "frameworks": fw_lookup,
        "categories": categories,
        "statistics": stats,
    }
    with open(metadata_file, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    print(f"  Saved {metadata_file}")

    print("\nDone!")


if __name__ == "__main__":
    main()
